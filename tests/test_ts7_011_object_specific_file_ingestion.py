"""TS7-011 CSV/XLSX staging for object-specific revisions."""

import io
import json
import os
import tempfile
import unittest
import uuid
import zipfile
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.auth import hash_password
from app.main import create_app
from app.persistence import AnalystStore
from tests.auth_test_helpers import csrf_headers, login_json_with_csrf, put_json_with_csrf
from tests.test_ts7_010_object_specific_series import DEFINITION


POSTGRES_TEST_DATABASE_URL = os.environ.get("POSTGRES_TEST_DATABASE_URL")


CSV_CONTENT = (
    "timestamp,duration_hours,value_mwh,quality\n"
    "2026-08-31T00:00:00-04:00,1,18.4,forecast\n"
    "2026-08-31T01:00:00-04:00,1,19.1,measured\n"
)

CSV_MAPPING = {
    "mode": "replace_full",
    "expected_base": None,
    "revision_contract": {
        "data_class_key": "forecast",
        "timezone": "America/Santiago",
        "regularity": "regular",
        "nominal_resolution_seconds": 3600,
    },
    "columns": {
        "timestamp_start": "timestamp",
        "timestamp_end": None,
        "duration_hours": "duration_hours",
        "signals": [
            {
                "series_key": "local_price_forecast",
                "value": "value_mwh",
                "quality_flag": "quality",
            }
        ],
    },
    "source": {
        "kind": "csv",
        "display_name": "Pronostico del proveedor",
        "external_reference": "forecast-2026-08-30",
    },
}


def make_xlsx_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Afluentes"
    sheet.append(["timestamp", "duration_hours", "value_mwh", "quality"])
    sheet.append(["2026-08-31T00:00:00-04:00", 1, 18.4, "forecast"])
    sheet.append(["2026-08-31T01:00:00-04:00", 1, 19.1, "measured"])
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def make_multi_sheet_xlsx_bytes() -> bytes:
    workbook = Workbook()
    first = workbook.active
    first.title = "Resumen"
    first.append(["timestamp", "duration_hours", "value_mwh", "quality"])
    first.append(["2026-08-31T00:00:00-04:00", 1, 999.0, "suspect"])
    selected = workbook.create_sheet("Afluentes")
    selected.append(["timestamp", "duration_hours", "value_mwh", "quality"])
    selected.append(["2026-08-31T00:00:00-04:00", 1, 18.4, "forecast"])
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def make_formula_xlsx_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Afluentes"
    sheet.append(["timestamp", "duration_hours", "value_mwh", "quality"])
    sheet.append(["2026-08-31T00:00:00-04:00", 1, "=20-1.6", "forecast"])
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def make_xlsx_compression_bomb() -> bytes:
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("xl/worksheets/sheet1.xml", "0" * 100_000)
    return output.getvalue()


class ObjectSpecificFileIngestionApiTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.store = AnalystStore("sqlite:///:memory:")
        self.client = TestClient(
            create_app(
                store=self.store,
                auth_enabled=True,
                input_source_root=Path(self.temp_dir.name) / "input-sources",
            )
        )
        self.store.create_user(
            email="analyst@example.local",
            display_name="Analyst",
            role="analyst",
            password_hash=hash_password("analyst pass"),
        )
        login = login_json_with_csrf(
            self.client, "analyst@example.local", "analyst pass"
        )
        self.assertEqual(login.status_code, 200, login.text)
        self.project = self.store.create_project(name="Cuenca Norte")
        self.object = self.store.ensure_global_signal_slot(
            project_id=self.project["id"], display_name="Sistema"
        )
        created = self.client.post(
            f"{self.root}/object-series",
            json={
                **DEFINITION,
                "source_expectation": {
                    "kind": "csv",
                    "display_name": "Pronostico del proveedor",
                },
            },
            headers={
                **csrf_headers(self.client),
                "Idempotency-Key": f"define-{uuid.uuid4().hex}",
            },
        )
        self.assertEqual(created.status_code, 201, created.text)
        self.signal_id = created.json()["object_series"]["signal_id"]
        self.definition_etag = created.headers["etag"]

    def tearDown(self):
        self.store.close()
        self.temp_dir.cleanup()

    @property
    def root(self):
        return (
            f"/api/projects/{self.project['id']}/linkable-objects/"
            f"{self.object['id']}/time-series"
        )

    @property
    def target(self):
        return f"{self.root}/object-series/{self.signal_id}"

    def upload_csv(self, *, content=CSV_CONTENT, data=None, key=None):
        return self.client.post(
            f"{self.target}/revision-ingestions/files",
            files={"file": ("forecast.csv", content, "text/csv")},
            data={} if data is None else data,
            headers={
                **csrf_headers(self.client),
                "Idempotency-Key": key or f"upload-{uuid.uuid4().hex}",
            },
        )

    def upload_xlsx(self, *, data, key=None, content=None):
        return self.client.post(
            f"{self.target}/revision-ingestions/files",
            files={
                "file": (
                    "forecast.xlsx",
                    make_xlsx_bytes() if content is None else content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            data=data,
            headers={
                **csrf_headers(self.client),
                "Idempotency-Key": key or f"upload-{uuid.uuid4().hex}",
            },
        )

    def test_csv_without_mapping_creates_a_safe_awaiting_mapping_job(self):
        uploaded = self.upload_csv()

        self.assertEqual(uploaded.status_code, 202, uploaded.text)
        ingestion = uploaded.json()["ingestion"]
        self.assertTrue(ingestion["ingestion_id"].startswith("tsi_"))
        self.assertEqual(ingestion["channel"], "file_csv")
        self.assertEqual(ingestion["state"], "awaiting_mapping")
        self.assertFalse(ingestion["capabilities"]["publish"])

        self.assertTrue(ingestion["capabilities"]["remap"])
        self.assertEqual(
            ingestion["file"],
            {
                "original_filename": "forecast.csv",
                "media_type": "text/csv",
                "available_sheets": [],
                "selected_sheet": None,
                "columns": [
                    "timestamp",
                    "duration_hours",
                    "value_mwh",
                    "quality",
                ],
                "preview_rows": [
                    {
                        "timestamp": "2026-08-31T00:00:00-04:00",
                        "duration_hours": "1",
                        "value_mwh": "18.4",
                        "quality": "forecast",
                    },
                    {
                        "timestamp": "2026-08-31T01:00:00-04:00",
                        "duration_hours": "1",
                        "value_mwh": "19.1",
                        "quality": "measured",
                    },
                ],
                "mapping_suggestions": {
                    "timestamp_start": "timestamp",
                    "timestamp_end": None,
                    "duration_hours": "duration_hours",
                    "signals": [
                        {
                            "series_key": "local_price_forecast",
                            "value": None,
                            "quality_flag": "quality",
                        }
                    ],
                },
            },
        )
        self.assertNotIn("stored_path", uploaded.text)
        detail = self.client.get(self.target).json()["object_series"]
        self.assertIsNone(detail["current_revision"])
        self.assertEqual(detail["availability"], "awaiting_data")

    def test_csv_mapping_previews_and_publishes_the_exact_staged_snapshot(self):
        uploaded = self.upload_csv(data={"mapping": json.dumps(CSV_MAPPING)})

        self.assertEqual(uploaded.status_code, 202, uploaded.text)
        ingestion = uploaded.json()["ingestion"]
        self.assertEqual(ingestion["state"], "ready_to_publish")
        self.assertEqual(ingestion["normalized"]["period_count"], 2)
        self.assertEqual(ingestion["normalized"]["value_count"], 2)
        self.assertTrue(ingestion["validation"]["valid"])
        self.assertTrue(ingestion["capabilities"]["remap"])

        preview = self.client.get(
            f"{self.target}/revision-ingestions/{ingestion['ingestion_id']}"
            "/preview?max_rows=1"
        )
        self.assertEqual(preview.status_code, 200, preview.text)
        self.assertEqual(preview.json()["source_row_count"], 2)
        self.assertEqual(preview.json()["returned_row_count"], 1)
        self.assertEqual(
            preview.json()["rows"],
            [
                {
                    "period_index": 0,
                    "timestamp_start": "2026-08-31T04:00:00Z",
                    "timestamp_end": "2026-08-31T05:00:00Z",
                    "series_key": "local_price_forecast",
                    "value": 18.4,
                    "quality_flag": "forecast",
                }
            ],
        )

        published = self.client.post(
            f"{self.target}/revision-ingestions/{ingestion['ingestion_id']}"
            "/publications",
            json={
                "validation_token": ingestion["validation_token"],
                "confirm": False,
                "reason_code": "forecast_refresh",
            },
            headers={
                **csrf_headers(self.client),
                "If-Match": self.definition_etag,
                "Idempotency-Key": "publish-csv-01",
            },
        )

        self.assertEqual(published.status_code, 201, published.text)
        publication = published.json()["publication"]
        self.assertEqual(
            publication["content_hash"], ingestion["normalized"]["content_hash"]
        )
        self.assertEqual(publication["source"]["kind"], "csv")
        canonical = self.client.get(
            f"{self.target}/preview?revision_id={publication['revision_id']}"
            "&from=2026-08-31T00:00:00Z&to=2026-09-01T00:00:00Z&max_points=500"
        )
        self.assertEqual(canonical.status_code, 200, canonical.text)
        self.assertEqual(
            [point["value"] for point in canonical.json()["points"]], [18.4, 19.1]
        )

    def test_xlsx_selected_sheet_publishes_the_same_normalized_contract(self):
        mapping = {
            **CSV_MAPPING,
            "sheet_name": "Afluentes",
            "source": {
                **CSV_MAPPING["source"],
                "kind": "xlsx",
            },
        }
        uploaded = self.upload_xlsx(data={"mapping": json.dumps(mapping)})

        self.assertEqual(uploaded.status_code, 202, uploaded.text)
        ingestion = uploaded.json()["ingestion"]
        self.assertEqual(ingestion["channel"], "file_xlsx")
        self.assertEqual(ingestion["state"], "ready_to_publish")
        self.assertEqual(ingestion["file"]["available_sheets"], ["Afluentes"])
        self.assertEqual(ingestion["file"]["selected_sheet"], "Afluentes")

        published = self.client.post(
            f"{self.target}/revision-ingestions/{ingestion['ingestion_id']}"
            "/publications",
            json={
                "validation_token": ingestion["validation_token"],
                "confirm": False,
                "reason_code": "forecast_refresh",
            },
            headers={
                **csrf_headers(self.client),
                "If-Match": self.definition_etag,
                "Idempotency-Key": "publish-xlsx-01",
            },
        )

        self.assertEqual(published.status_code, 201, published.text)
        publication = published.json()["publication"]
        self.assertEqual(publication["revision_number"], 1)
        self.assertEqual(publication["source"]["kind"], "xlsx")
        self.assertEqual(
            publication["content_hash"], ingestion["normalized"]["content_hash"]
        )

    def test_a_second_file_load_creates_a_revision_without_changing_identity(self):
        first_ingestion = self.upload_csv(
            data={"mapping": json.dumps(CSV_MAPPING)}
        ).json()["ingestion"]
        first_response = self.client.post(
            f"{self.target}/revision-ingestions/{first_ingestion['ingestion_id']}"
            "/publications",
            json={
                "validation_token": first_ingestion["validation_token"],
                "confirm": False,
                "reason_code": "initial_file",
            },
            headers={
                **csrf_headers(self.client),
                "If-Match": self.definition_etag,
                "Idempotency-Key": "publish-file-first",
            },
        )
        self.assertEqual(first_response.status_code, 201, first_response.text)
        first = first_response.json()["publication"]
        after_first = self.client.get(self.target)

        replacement_mapping = {
            **CSV_MAPPING,
            "expected_base": {
                "revision_id": first["revision_id"],
                "content_hash": first["content_hash"],
            },
        }
        second_ingestion_response = self.upload_csv(
            content=CSV_CONTENT.replace("18.4", "21.0").replace("19.1", "22.0"),
            data={"mapping": json.dumps(replacement_mapping)},
        )
        self.assertEqual(
            second_ingestion_response.status_code,
            202,
            second_ingestion_response.text,
        )
        second_ingestion = second_ingestion_response.json()["ingestion"]
        self.assertEqual(second_ingestion["base"], replacement_mapping["expected_base"])
        second_response = self.client.post(
            f"{self.target}/revision-ingestions/{second_ingestion['ingestion_id']}"
            "/publications",
            json={
                "validation_token": second_ingestion["validation_token"],
                "confirm": False,
                "reason_code": "replacement_file",
            },
            headers={
                **csrf_headers(self.client),
                "If-Match": after_first.headers["etag"],
                "Idempotency-Key": "publish-file-second",
            },
        )

        self.assertEqual(second_response.status_code, 201, second_response.text)
        second = second_response.json()["publication"]
        self.assertEqual(second["revision_number"], 2)
        self.assertEqual(second["signal_ids"], [self.signal_id])
        self.assertEqual(second["set_id"], first["set_id"])
        history = self.client.get(f"{self.target}/revisions").json()
        self.assertEqual([item["number"] for item in history["items"]], [2, 1])

        old_preview = self.client.get(
            f"{self.target}/preview?revision_id={first['revision_id']}"
            "&from=2026-08-31T00:00:00Z&to=2026-09-01T00:00:00Z&max_points=500"
        )
        self.assertEqual(
            [point["value"] for point in old_preview.json()["points"]], [18.4, 19.1]
        )

    def test_file_mapping_can_be_set_then_corrected_without_another_upload(self):
        uploaded = self.upload_csv()
        self.assertEqual(uploaded.status_code, 202, uploaded.text)
        ingestion_id = uploaded.json()["ingestion"]["ingestion_id"]
        mapping_url = f"{self.target}/revision-ingestions/{ingestion_id}/mapping"
        wrong_mapping = {
            **CSV_MAPPING,
            "columns": {
                **CSV_MAPPING["columns"],
                "signals": [
                    {
                        "series_key": "local_price_forecast",
                        "value": "missing_value_column",
                        "quality_flag": "quality",
                    }
                ],
            },
        }

        refused = put_json_with_csrf(self.client, mapping_url, wrong_mapping)

        self.assertEqual(refused.status_code, 422, refused.text)
        self.assertEqual(refused.json()["code"], "TS_INGEST_MAPPING_INVALID")
        invalid_job = refused.json()["context"]["ingestion"]
        self.assertEqual(invalid_job["ingestion_id"], ingestion_id)
        self.assertEqual(invalid_job["state"], "awaiting_mapping")
        self.assertFalse(invalid_job["capabilities"]["publish"])

        corrected = put_json_with_csrf(self.client, mapping_url, CSV_MAPPING)

        self.assertEqual(corrected.status_code, 200, corrected.text)
        ready = corrected.json()["ingestion"]
        self.assertEqual(ready["ingestion_id"], ingestion_id)
        self.assertEqual(ready["state"], "ready_to_publish")
        self.assertTrue(ready["validation"]["valid"])
        self.assertTrue(ready["capabilities"]["publish"])
        preview = self.client.get(
            f"{self.target}/revision-ingestions/{ingestion_id}/preview?max_rows=2"
        )
        self.assertEqual(
            [row["value"] for row in preview.json()["rows"]], [18.4, 19.1]
        )

    def test_cancelling_a_file_job_retires_staging_and_leaves_no_revision(self):
        ingestion = self.upload_csv(
            data={"mapping": json.dumps(CSV_MAPPING)}
        ).json()["ingestion"]
        job_url = f"{self.target}/revision-ingestions/{ingestion['ingestion_id']}"

        cancelled = self.client.delete(
            job_url, headers=csrf_headers(self.client)
        )

        self.assertEqual(cancelled.status_code, 204, cancelled.text)
        status = self.client.get(job_url)
        self.assertEqual(status.status_code, 200, status.text)
        self.assertEqual(status.json()["ingestion"]["state"], "cancelled")
        self.assertEqual(status.json()["ingestion"]["file"]["preview_rows"], [])
        preview = self.client.get(f"{job_url}/preview")
        self.assertEqual(preview.status_code, 410, preview.text)
        self.assertEqual(preview.json()["code"], "TS_INGEST_SESSION_UNAVAILABLE")
        publication = self.client.post(
            f"{job_url}/publications",
            json={
                "validation_token": ingestion["validation_token"],
                "confirm": False,
                "reason_code": "must_not_publish",
            },
            headers={
                **csrf_headers(self.client),
                "If-Match": self.definition_etag,
                "Idempotency-Key": "publish-cancelled-file",
            },
        )
        self.assertEqual(publication.status_code, 410, publication.text)
        detail = self.client.get(self.target).json()["object_series"]
        self.assertIsNone(detail["current_revision"])

    def test_file_upload_idempotency_replays_and_refuses_key_reuse(self):
        missing_key = self.client.post(
            f"{self.target}/revision-ingestions/files",
            files={"file": ("forecast.csv", CSV_CONTENT, "text/csv")},
            headers=csrf_headers(self.client),
        )
        self.assertEqual(missing_key.status_code, 428, missing_key.text)
        self.assertEqual(
            missing_key.json()["code"], "TS_INGEST_PRECONDITION_REQUIRED"
        )

        first = self.upload_csv(key="stable-file-upload")
        replay = self.upload_csv(key="stable-file-upload")

        self.assertEqual(first.status_code, 202, first.text)
        self.assertEqual(replay.status_code, 202, replay.text)
        self.assertEqual(
            replay.json()["ingestion"]["ingestion_id"],
            first.json()["ingestion"]["ingestion_id"],
        )

        conflict = self.upload_csv(
            content=CSV_CONTENT.replace("18.4", "77.0"),
            key="stable-file-upload",
        )
        self.assertEqual(conflict.status_code, 409, conflict.text)
        self.assertEqual(
            conflict.json()["code"], "TS_INGEST_IDEMPOTENCY_CONFLICT"
        )

    def test_a_file_over_the_column_quota_is_invalid_and_writes_no_revision(self):
        columns = [f"column_{index}" for index in range(201)]
        oversized = ",".join(columns) + "\n" + ",".join("1" for _ in columns) + "\n"

        uploaded = self.upload_csv(content=oversized)

        self.assertEqual(uploaded.status_code, 202, uploaded.text)
        ingestion = uploaded.json()["ingestion"]
        self.assertEqual(ingestion["state"], "invalid")
        self.assertEqual(
            ingestion["validation"]["errors"][0]["code"],
            "TS_INGEST_QUOTA_EXCEEDED",
        )
        self.assertFalse(ingestion["capabilities"]["publish"])
        detail = self.client.get(self.target).json()["object_series"]
        self.assertIsNone(detail["current_revision"])

    def test_file_append_tail_publishes_a_complete_snapshot(self):
        first_ingestion = self.upload_csv(
            data={"mapping": json.dumps(CSV_MAPPING)}
        ).json()["ingestion"]
        first_response = self.client.post(
            f"{self.target}/revision-ingestions/{first_ingestion['ingestion_id']}"
            "/publications",
            json={
                "validation_token": first_ingestion["validation_token"],
                "confirm": False,
                "reason_code": "initial_file",
            },
            headers={
                **csrf_headers(self.client),
                "If-Match": self.definition_etag,
                "Idempotency-Key": "publish-before-file-tail",
            },
        )
        self.assertEqual(first_response.status_code, 201, first_response.text)
        first = first_response.json()["publication"]
        current = self.client.get(self.target)
        tail_mapping = {
            **CSV_MAPPING,
            "mode": "append_tail",
            "expected_base": {
                "revision_id": first["revision_id"],
                "content_hash": first["content_hash"],
            },
        }
        tail_csv = (
            "timestamp,duration_hours,value_mwh,quality\n"
            "2026-08-31T02:00:00-04:00,1,20.0,forecast\n"
            "2026-08-31T03:00:00-04:00,1,21.0,forecast\n"
        )

        staged_response = self.upload_csv(
            content=tail_csv, data={"mapping": json.dumps(tail_mapping)}
        )
        self.assertEqual(staged_response.status_code, 202, staged_response.text)
        staged = staged_response.json()["ingestion"]
        self.assertEqual(staged["state"], "ready_to_publish")
        self.assertEqual(staged["normalized"]["period_count"], 4)
        publication_response = self.client.post(
            f"{self.target}/revision-ingestions/{staged['ingestion_id']}/publications",
            json={
                "validation_token": staged["validation_token"],
                "confirm": False,
                "reason_code": "append_file_tail",
            },
            headers={
                **csrf_headers(self.client),
                "If-Match": current.headers["etag"],
                "Idempotency-Key": "publish-file-tail",
            },
        )

        self.assertEqual(publication_response.status_code, 201, publication_response.text)
        publication = publication_response.json()["publication"]
        preview = self.client.get(
            f"{self.target}/preview?revision_id={publication['revision_id']}"
            "&from=2026-08-31T00:00:00Z&to=2026-09-01T00:00:00Z&max_points=500"
        )
        self.assertEqual(
            [point["value"] for point in preview.json()["points"]],
            [18.4, 19.1, 20.0, 21.0],
        )

    def test_multi_sheet_xlsx_requires_then_revalidates_an_explicit_sheet(self):
        uploaded = self.upload_xlsx(
            content=make_multi_sheet_xlsx_bytes(), data={}
        )
        self.assertEqual(uploaded.status_code, 202, uploaded.text)
        ingestion = uploaded.json()["ingestion"]
        self.assertEqual(ingestion["state"], "awaiting_mapping")
        self.assertEqual(
            ingestion["file"]["available_sheets"], ["Resumen", "Afluentes"]
        )
        mapping_url = (
            f"{self.target}/revision-ingestions/{ingestion['ingestion_id']}/mapping"
        )

        missing_sheet = put_json_with_csrf(
            self.client,
            mapping_url,
            {
                **CSV_MAPPING,
                "source": {**CSV_MAPPING["source"], "kind": "xlsx"},
            },
        )

        self.assertEqual(missing_sheet.status_code, 422, missing_sheet.text)
        self.assertEqual(missing_sheet.json()["code"], "TS_INGEST_MAPPING_INVALID")
        with_sheet = {
            **CSV_MAPPING,
            "sheet_name": "Afluentes",
            "source": {**CSV_MAPPING["source"], "kind": "xlsx"},
        }
        selected = put_json_with_csrf(self.client, mapping_url, with_sheet)
        self.assertEqual(selected.status_code, 200, selected.text)
        ready = selected.json()["ingestion"]
        self.assertEqual(ready["state"], "ready_to_publish")
        self.assertEqual(ready["file"]["selected_sheet"], "Afluentes")
        preview = self.client.get(
            f"{self.target}/revision-ingestions/{ingestion['ingestion_id']}/preview"
        )
        self.assertEqual([row["value"] for row in preview.json()["rows"]], [18.4])

    def test_a_formula_xlsx_is_refused_with_no_partial_revision(self):
        mapping = {
            **CSV_MAPPING,
            "sheet_name": "Afluentes",
            "source": {**CSV_MAPPING["source"], "kind": "xlsx"},
        }

        refused = self.upload_xlsx(
            content=make_formula_xlsx_bytes(),
            data={"mapping": json.dumps(mapping)},
        )

        self.assertEqual(refused.status_code, 422, refused.text)
        self.assertEqual(
            refused.headers["content-type"].split(";")[0],
            "application/problem+json",
        )
        self.assertEqual(refused.json()["code"], "TS_INGEST_VALIDATION_FAILED")
        self.assertIn("formulas", refused.json()["detail"])
        self.assertNotIn("ingestion", refused.json())
        detail = self.client.get(self.target).json()["object_series"]
        self.assertIsNone(detail["current_revision"])

    def test_file_validation_errors_use_file_locations_not_json_pointers(self):
        half_hour_contract = {
            **CSV_MAPPING,
            "revision_contract": {
                **CSV_MAPPING["revision_contract"],
                "nominal_resolution_seconds": 1800,
            },
        }

        uploaded = self.upload_csv(
            data={"mapping": json.dumps(half_hour_contract)}
        )

        self.assertEqual(uploaded.status_code, 202, uploaded.text)
        ingestion = uploaded.json()["ingestion"]
        self.assertEqual(ingestion["state"], "invalid")
        error = ingestion["validation"]["errors"][0]
        self.assertEqual(error["code"], "TS_INGEST_TEMPORAL_CONTRACT_INVALID")
        self.assertEqual(
            error["location"],
            {
                "record_index": 0,
                "source_row_number": 2,
                "column": "duration_hours",
            },
        )

    def test_only_three_file_jobs_can_be_active_per_actor_and_project(self):
        jobs = [self.upload_csv(key=f"quota-upload-{index}") for index in range(3)]
        self.assertEqual([response.status_code for response in jobs], [202, 202, 202])

        refused = self.upload_csv(key="quota-upload-four")

        self.assertEqual(refused.status_code, 422, refused.text)
        self.assertEqual(refused.json()["code"], "TS_INGEST_QUOTA_EXCEEDED")
        first_job = jobs[0].json()["ingestion"]["ingestion_id"]
        cancelled = self.client.delete(
            f"{self.target}/revision-ingestions/{first_job}",
            headers=csrf_headers(self.client),
        )
        self.assertEqual(cancelled.status_code, 204, cancelled.text)
        admitted = self.upload_csv(key="quota-upload-four")
        self.assertEqual(admitted.status_code, 202, admitted.text)

    def test_xlsx_compression_ratio_is_refused_before_parsing(self):
        refused = self.upload_xlsx(
            content=make_xlsx_compression_bomb(), data={}
        )

        self.assertEqual(refused.status_code, 413, refused.text)
        self.assertEqual(refused.json()["code"], "TS_INGEST_PAYLOAD_TOO_LARGE")
        self.assertIsNone(
            self.client.get(self.target).json()["object_series"]["current_revision"]
        )

    def test_file_source_cannot_supply_server_owned_provenance(self):
        forged = {
            **CSV_MAPPING,
            "source": {
                **CSV_MAPPING["source"],
                "stored_path": "C:/secrets/source.csv",
                "checksum": "sha256:forged",
                "created_by": "admin@example.local",
            },
        }

        refused = self.upload_csv(data={"mapping": json.dumps(forged)})

        self.assertEqual(refused.status_code, 422, refused.text)
        self.assertEqual(refused.json()["code"], "TS_INGEST_MAPPING_INVALID")
        self.assertIsNone(
            self.client.get(self.target).json()["object_series"]["current_revision"]
        )

    def test_duplicate_signal_mappings_are_never_collapsed_silently(self):
        duplicate = {
            **CSV_MAPPING,
            "columns": {
                **CSV_MAPPING["columns"],
                "signals": [
                    CSV_MAPPING["columns"]["signals"][0],
                    {
                        **CSV_MAPPING["columns"]["signals"][0],
                        "value": "quality",
                    },
                ],
            },
        }

        uploaded = self.upload_csv(data={"mapping": json.dumps(duplicate)})

        self.assertEqual(uploaded.status_code, 202, uploaded.text)
        ingestion = uploaded.json()["ingestion"]
        self.assertEqual(ingestion["state"], "awaiting_mapping")
        self.assertEqual(
            ingestion["validation"]["errors"][0]["code"],
            "TS_INGEST_MAPPING_INVALID",
        )
        self.assertFalse(ingestion["capabilities"]["publish"])


@unittest.skipUnless(
    POSTGRES_TEST_DATABASE_URL,
    "set POSTGRES_TEST_DATABASE_URL to run the PostgreSQL contract tests",
)
class PostgresObjectSpecificFileIngestionApiTests(
    ObjectSpecificFileIngestionApiTests
):
    """Mirror the object file-ingestion HTTP contract on PostgreSQL."""

    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.store = AnalystStore(POSTGRES_TEST_DATABASE_URL)
        self.rollback = self.store.connection._connection.transaction(  # noqa: SLF001
            force_rollback=True
        )
        self.rollback.__enter__()
        suffix = uuid.uuid4().hex[:10]
        self.client = TestClient(
            create_app(
                store=self.store,
                auth_enabled=True,
                input_source_root=Path(self.temp_dir.name) / "input-sources",
            )
        )
        email = f"ts7-011-{suffix}@example.local"
        self.store.create_user(
            email=email,
            display_name="PostgreSQL Analyst",
            role="analyst",
            password_hash=hash_password("postgres analyst pass"),
        )
        login = login_json_with_csrf(
            self.client, email, "postgres analyst pass"
        )
        self.assertEqual(login.status_code, 200, login.text)
        self.project = self.store.create_project(name=f"Cuenca Norte {suffix}")
        self.object = self.store.ensure_global_signal_slot(
            project_id=self.project["id"], display_name="Sistema"
        )
        created = self.client.post(
            f"{self.root}/object-series",
            json={
                **DEFINITION,
                "source_expectation": {
                    "kind": "csv",
                    "display_name": "Pronostico del proveedor",
                },
            },
            headers={
                **csrf_headers(self.client),
                "Idempotency-Key": f"define-{uuid.uuid4().hex}",
            },
        )
        self.assertEqual(created.status_code, 201, created.text)
        self.signal_id = created.json()["object_series"]["signal_id"]
        self.definition_etag = created.headers["etag"]

    def tearDown(self):
        try:
            self.rollback.__exit__(None, None, None)
        finally:
            self.store.close()
            self.temp_dir.cleanup()


if __name__ == "__main__":
    unittest.main()
