from __future__ import annotations

import copy
import csv
import hashlib
import io
import math
import re
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


class TimeSeriesIngestionError(ValueError):
    pass


XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def ingest_time_series_source(
    *,
    draft_document: dict[str, Any],
    original_filename: str,
    content_type: str | None,
    content: bytes,
    input_source_root: Path,
    preview_limit: int = 5,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    safe_filename = safe_source_filename(original_filename)
    if is_xlsx_source(safe_filename, content_type):
        return ingest_xlsx_source(
            draft_document=draft_document,
            original_filename=safe_filename,
            content_type=content_type,
            content=content,
            input_source_root=input_source_root,
            preview_limit=preview_limit,
            sheet_name=sheet_name,
        )
    return ingest_csv_source(
        draft_document=draft_document,
        original_filename=safe_filename,
        content_type=content_type,
        content=content,
        input_source_root=input_source_root,
        preview_limit=preview_limit,
    )


def ingest_csv_source(
    *,
    draft_document: dict[str, Any],
    original_filename: str,
    content_type: str | None,
    content: bytes,
    input_source_root: Path,
    preview_limit: int = 5,
) -> dict[str, Any]:
    safe_filename = safe_source_filename(original_filename)
    source_id = f"csv_{uuid.uuid4().hex[:12]}"
    stored_path = input_source_root / f"{source_id}_{safe_filename}"
    stored_path.parent.mkdir(parents=True, exist_ok=True)
    stored_path.write_bytes(content)

    text = decode_csv_content(content)
    columns, rows = parse_csv_preview(text, preview_limit=preview_limit)
    mapping_suggestions = suggest_mappings(columns, draft_document)

    return {
        "id": source_id,
        "kind": "csv",
        "original_filename": safe_filename,
        "media_type": content_type or "text/csv",
        "checksum": source_checksum(content),
        "stored_path": str(stored_path),
        "columns": columns,
        "preview_rows": rows,
        "mapping_suggestions": mapping_suggestions,
        "mapping": copy.deepcopy(mapping_suggestions),
    }


def ingest_xlsx_source(
    *,
    draft_document: dict[str, Any],
    original_filename: str,
    content_type: str | None,
    content: bytes,
    input_source_root: Path,
    preview_limit: int = 5,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    safe_filename = safe_source_filename(original_filename)
    source_id = f"xlsx_{uuid.uuid4().hex[:12]}"
    stored_path = input_source_root / f"{source_id}_{safe_filename}"
    stored_path.parent.mkdir(parents=True, exist_ok=True)
    stored_path.write_bytes(content)

    selected_sheet, columns, rows = parse_xlsx_preview(
        content,
        preview_limit=preview_limit,
        sheet_name=sheet_name,
    )
    mapping_suggestions = suggest_mappings(columns, draft_document)

    return {
        "id": source_id,
        "kind": "xlsx",
        "original_filename": safe_filename,
        "media_type": content_type or XLSX_MEDIA_TYPE,
        "checksum": source_checksum(content),
        "stored_path": str(stored_path),
        "selected_sheet": selected_sheet,
        "columns": columns,
        "preview_rows": rows,
        "mapping_suggestions": mapping_suggestions,
        "mapping": copy.deepcopy(mapping_suggestions),
    }


def attach_time_series_source(
    document: dict[str, Any],
    source: dict[str, Any],
) -> dict[str, Any]:
    updated = copy.deepcopy(document)
    time_series = updated.get("time_series")
    if not isinstance(time_series, dict):
        time_series = {}
        updated["time_series"] = time_series
    existing_sources = time_series.get("sources")
    if not isinstance(existing_sources, list):
        existing_sources = []
    time_series["sources"] = [item for item in existing_sources if not isinstance(item, dict) or item.get("id") != source["id"]]
    time_series["sources"].append(source)
    time_series["active_source_id"] = source["id"]
    return updated


def apply_time_series_mapping(
    *,
    document: dict[str, Any],
    source_id: str,
    mapping: dict[str, Any],
    input_source_root: Path,
) -> tuple[dict[str, Any], dict[str, Any]]:
    updated = copy.deepcopy(document)
    source = find_source(updated, source_id)
    columns, rows = read_time_series_source_rows(source, input_source_root)
    validation, validated_rows = validate_mapping(
        columns=columns,
        rows=rows,
        mapping=mapping,
        draft_document=updated,
    )
    source["mapping"] = copy.deepcopy(mapping)
    source["validation"] = validation
    source["validated_rows"] = validated_rows
    return updated, copy.deepcopy(source)


def get_time_series_source_rows(
    *,
    document: dict[str, Any],
    source_id: str,
    input_source_root: Path,
) -> tuple[list[str], list[dict[str, str]]]:
    source = find_source(document, source_id)
    columns, rows = read_time_series_source_rows(source, input_source_root)
    return list(columns), copy.deepcopy(rows)


def update_time_series_source_rows(
    *,
    document: dict[str, Any],
    source_id: str,
    rows: list[dict[str, Any]],
    input_source_root: Path,
) -> tuple[dict[str, Any], dict[str, Any]]:
    updated = copy.deepcopy(document)
    source = find_source(updated, source_id)
    columns, current_rows = read_time_series_source_rows(source, input_source_root)
    normalized_rows = normalize_edited_rows(rows, columns)
    if len(normalized_rows) != len(current_rows):
        raise TimeSeriesIngestionError("edited time-series data must preserve the original row count")

    source["edited_rows"] = normalized_rows
    source["preview_rows"] = copy.deepcopy(normalized_rows[:5])

    mapping = source.get("mapping")
    if isinstance(mapping, dict):
        validation, validated_rows = validate_mapping(
            columns=columns,
            rows=normalized_rows,
            mapping=mapping,
            draft_document=updated,
        )
        source["validation"] = validation
        source["validated_rows"] = validated_rows
    else:
        source.pop("validation", None)
        source.pop("validated_rows", None)
    return updated, copy.deepcopy(source)


def read_time_series_source_rows(
    source: dict[str, Any],
    input_source_root: Path,
) -> tuple[list[str], list[dict[str, str]]]:
    columns = source.get("columns")
    edited_rows = source.get("edited_rows")
    if isinstance(columns, list) and isinstance(edited_rows, list):
        return [str(column) for column in columns], normalize_edited_rows(edited_rows, [str(column) for column in columns])

    stored_path = safe_stored_source_path(source, input_source_root)
    if source.get("kind") == "xlsx":
        _, parsed_columns, rows = parse_xlsx_rows(
            stored_path.read_bytes(),
            sheet_name=source.get("selected_sheet"),
        )
    else:
        text = decode_csv_content(stored_path.read_bytes())
        parsed_columns, rows = parse_csv_rows(text)
    return parsed_columns, rows


def normalize_edited_rows(rows: list[dict[str, Any]], columns: list[str]) -> list[dict[str, str]]:
    if not isinstance(rows, list):
        raise TimeSeriesIngestionError("edited time-series data must be a list of rows")

    normalized_rows: list[dict[str, str]] = []
    allowed_columns = set(columns)
    for row_index, row in enumerate(rows, start=2):
        if not isinstance(row, dict):
            raise TimeSeriesIngestionError(f"row {row_index}: edited row must be an object")
        extra_columns = set(row) - allowed_columns
        if extra_columns:
            extras = ", ".join(sorted(str(column) for column in extra_columns))
            raise TimeSeriesIngestionError(f"row {row_index}: unknown columns: {extras}")

        normalized_row: dict[str, str] = {}
        for column in columns:
            value = row.get(column)
            if value is None:
                normalized_row[column] = ""
            elif isinstance(value, (str, int, float, bool)):
                normalized_row[column] = str(value)
            else:
                raise TimeSeriesIngestionError(f"row {row_index}: {column} must be a scalar value")
        normalized_rows.append(normalized_row)
    return normalized_rows


def find_source(document: dict[str, Any], source_id: str) -> dict[str, Any]:
    time_series = document.get("time_series")
    sources = time_series.get("sources") if isinstance(time_series, dict) else None
    if not isinstance(sources, list):
        raise KeyError(f"time-series source {source_id} not found")
    for source in sources:
        if isinstance(source, dict) and source.get("id") == source_id:
            return source
    raise KeyError(f"time-series source {source_id} not found")


def safe_stored_source_path(source: dict[str, Any], input_source_root: Path) -> Path:
    root = input_source_root.resolve(strict=False)
    path = Path(str(source.get("stored_path") or "")).resolve(strict=False)
    try:
        path.relative_to(root)
    except ValueError as error:
        raise TimeSeriesIngestionError("time-series source file path is outside the configured source root") from error
    if not path.is_file():
        raise TimeSeriesIngestionError("time-series source file is missing")
    return path


def is_xlsx_source(filename: str, content_type: str | None) -> bool:
    return filename.lower().endswith(".xlsx") or (content_type or "").lower() == XLSX_MEDIA_TYPE


def safe_source_filename(filename: str) -> str:
    basename = Path(filename or "source.csv").name.strip()
    if not basename:
        basename = "source.csv"
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", basename)
    return cleaned or "source.csv"


def decode_csv_content(content: bytes) -> str:
    try:
        return content.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise TimeSeriesIngestionError("CSV source file must be UTF-8 encoded") from error


def source_checksum(content: bytes) -> str:
    return "sha256:" + hashlib.sha256(content).hexdigest()


def parse_csv_preview(text: str, *, preview_limit: int) -> tuple[list[str], list[dict[str, str]]]:
    columns, rows = parse_csv_rows(text)
    return columns, rows[:preview_limit]


def parse_csv_rows(text: str) -> tuple[list[str], list[dict[str, str]]]:
    reader = csv.DictReader(io.StringIO(text))
    columns = list(reader.fieldnames or [])
    if not columns:
        raise TimeSeriesIngestionError("CSV source file must include a header row")

    rows: list[dict[str, str]] = []
    for row in reader:
        rows.append({column: str(row.get(column) or "") for column in columns})
    return columns, rows


def parse_xlsx_preview(
    content: bytes,
    *,
    preview_limit: int,
    sheet_name: str | None = None,
) -> tuple[str, list[str], list[dict[str, str]]]:
    selected_sheet, columns, rows = parse_xlsx_rows(content, sheet_name=sheet_name)
    return selected_sheet, columns, rows[:preview_limit]


def parse_xlsx_rows(content: bytes, *, sheet_name: str | None = None) -> tuple[str, list[str], list[dict[str, str]]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=False, data_only=False)
    except (InvalidFileException, OSError, KeyError, ValueError) as error:
        raise TimeSeriesIngestionError("XLSX source file could not be read") from error

    if not workbook.sheetnames:
        raise TimeSeriesIngestionError("XLSX source file must include at least one sheet")
    requested_sheet = str(sheet_name or "").strip()
    selected_sheet = requested_sheet or workbook.sheetnames[0]
    if selected_sheet not in workbook.sheetnames:
        raise TimeSeriesIngestionError(f"XLSX sheet {selected_sheet!r} was not found")

    sheet = workbook[selected_sheet]
    if sheet.merged_cells.ranges:
        raise TimeSeriesIngestionError("XLSX source file uses merged cells, which are not supported")
    if sheet.tables:
        raise TimeSeriesIngestionError("XLSX source file uses Excel tables, which are not supported")

    parsed_rows: list[list[str]] = []
    for row in sheet.iter_rows():
        values = [xlsx_cell_to_text(cell) for cell in row]
        parsed_rows.append(values)

    while parsed_rows and all(value == "" for value in parsed_rows[-1]):
        parsed_rows.pop()
    if not parsed_rows:
        raise TimeSeriesIngestionError("XLSX source file must include a header row")

    header = parsed_rows[0]
    while header and header[-1] == "":
        header.pop()
    columns = [str(value).strip() for value in header]
    if not columns or any(column == "" for column in columns):
        raise TimeSeriesIngestionError("XLSX source file must include nonempty column headers")
    if len(set(columns)) != len(columns):
        raise TimeSeriesIngestionError("XLSX source file contains duplicate column headers")

    rows: list[dict[str, str]] = []
    for raw_row in parsed_rows[1:]:
        values = raw_row[: len(columns)]
        values.extend([""] * (len(columns) - len(values)))
        rows.append({column: values[index] for index, column in enumerate(columns)})
    return selected_sheet, columns, rows


def xlsx_cell_to_text(cell) -> str:
    if cell.data_type == "f":
        raise TimeSeriesIngestionError("XLSX source file contains formulas, which are not supported")
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def validate_mapping(
    *,
    columns: list[str],
    rows: list[dict[str, str]],
    mapping: dict[str, Any],
    draft_document: dict[str, Any],
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    errors: list[str] = []
    required_columns = required_mapping_columns(mapping, draft_document, errors)
    for label, column in required_columns:
        if column not in columns:
            errors.append(f"{label} mapping references missing column {column!r}")
    if errors:
        return validation_error("mapping", errors), []

    validated_rows: list[dict[str, Any]] = []
    seen_timestamps: set[datetime] = set()
    previous_timestamp: datetime | None = None
    for row_index, row in enumerate(rows, start=2):
        normalized_row: dict[str, Any] = {}
        timestamp_value = str(row.get(mapping["timestamp"]) or "").strip()
        try:
            timestamp = parse_timestamp(timestamp_value)
        except ValueError:
            errors.append(f"row {row_index}: timestamp must be ISO-8601")
            continue
        if timestamp in seen_timestamps:
            errors.append(f"row {row_index}: duplicate timestamp {timestamp_value}")
        if previous_timestamp is not None and timestamp < previous_timestamp:
            errors.append(f"row {row_index}: timestamps must be sorted ascending")
        seen_timestamps.add(timestamp)
        previous_timestamp = timestamp
        normalized_row["timestamp"] = timestamp_value

        duration = parse_mapped_float(row, mapping["duration_hours"], row_index, "duration_hours", errors)
        if duration is not None and duration <= 0:
            errors.append(f"row {row_index}: duration_hours must be positive")
        normalized_row["duration_hours"] = duration

        legacy_price_column = mapped_column(mapping, "price_usd_per_mwh")
        import_price_column = mapped_column(mapping, "import_price_usd_per_mwh")
        export_price_column = mapped_column(mapping, "export_price_usd_per_mwh")
        normalized_row["price_usd_per_mwh"] = parse_mapped_float_or_zero(
            row,
            legacy_price_column,
            row_index,
            "price_usd_per_mwh",
            errors,
        )
        if import_price_column or export_price_column:
            normalized_row["import_price_usd_per_mwh"] = parse_mapped_float_or_zero(
                row,
                import_price_column,
                row_index,
                "import_price_usd_per_mwh",
                errors,
            )
            normalized_row["export_price_usd_per_mwh"] = parse_mapped_float_or_zero(
                row,
                export_price_column,
                row_index,
                "export_price_usd_per_mwh",
                errors,
            )

        renewable_values: dict[str, float | None] = {}
        for asset_id, column in mapped_asset_columns(
            mapping,
            draft_document,
            "renewable",
            "renewable_available_power_mw",
        ):
            value = parse_mapped_float_or_zero(row, column, row_index, f"renewable {asset_id}", errors)
            if value is not None and value < 0:
                errors.append(f"row {row_index}: renewable {asset_id} availability must be nonnegative")
            renewable_values[str(asset_id)] = value
        normalized_row["renewable_available_power_mw"] = renewable_values

        load_values: dict[str, float | None] = {}
        for asset_id, column in mapped_asset_columns(mapping, draft_document, "load", "load_demand_mw"):
            value = parse_mapped_float_or_zero(row, column, row_index, f"load {asset_id}", errors)
            if value is not None and value < 0:
                errors.append(f"row {row_index}: load {asset_id} demand must be nonnegative")
            load_values[str(asset_id)] = value
        normalized_row["load_demand_mw"] = load_values

        hydro_values: dict[str, float | None] = {}
        for asset_id, column in mapped_asset_columns(mapping, draft_document, "hydro", "hydro_inflow_m3s"):
            value = parse_mapped_float_or_zero(row, column, row_index, f"hydro {asset_id} inflow", errors)
            if value is not None and value < 0:
                errors.append(f"row {row_index}: hydro {asset_id} inflow must be nonnegative")
            hydro_values[str(asset_id)] = value
        normalized_row["hydro_inflow_m3s"] = hydro_values

        validated_rows.append(normalized_row)

    if errors:
        return validation_error("python_validation", errors), []
    return {"ok": True, "errors": []}, validated_rows


def validation_error(error_category: str, errors: list[str]) -> dict[str, Any]:
    return {
        "ok": False,
        "error_category": error_category,
        "errors": errors,
    }


def required_mapping_columns(
    mapping: dict[str, Any],
    draft_document: dict[str, Any],
    errors: list[str],
) -> list[tuple[str, str]]:
    required: list[tuple[str, str]] = []
    for key in ["timestamp", "duration_hours"]:
        column = mapping.get(key)
        if not column:
            errors.append(f"{key} mapping is required")
        else:
            required.append((key, str(column)))

    legacy_price = mapped_column(mapping, "price_usd_per_mwh")
    import_price = mapped_column(mapping, "import_price_usd_per_mwh")
    export_price = mapped_column(mapping, "export_price_usd_per_mwh")
    if legacy_price:
        required.append(("price_usd_per_mwh", str(legacy_price)))
    elif import_price or export_price:
        if import_price:
            required.append(("import_price_usd_per_mwh", str(import_price)))
        if export_price:
            required.append(("export_price_usd_per_mwh", str(export_price)))

    for asset_id, column in mapped_asset_columns(
        mapping,
        draft_document,
        "renewable",
        "renewable_available_power_mw",
    ):
        if column:
            required.append((f"renewable_available_power_mw.{asset_id}", str(column)))

    for asset_id, column in mapped_asset_columns(mapping, draft_document, "load", "load_demand_mw"):
        if column:
            required.append((f"load_demand_mw.{asset_id}", str(column)))

    for asset_id, column in mapped_asset_columns(mapping, draft_document, "hydro", "hydro_inflow_m3s"):
        if column:
            required.append((f"hydro_inflow_m3s.{asset_id}", str(column)))

    return required


def mapped_column(mapping: dict[str, Any], key: str) -> str | None:
    value = mapping.get(key)
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def mapped_asset_columns(
    mapping: dict[str, Any],
    draft_document: dict[str, Any],
    asset_type: str,
    mapping_key: str,
) -> list[tuple[str, str | None]]:
    raw_asset_mapping = mapping.get(mapping_key)
    if isinstance(raw_asset_mapping, dict):
        asset_mapping = {str(asset_id): column for asset_id, column in raw_asset_mapping.items()}
    else:
        asset_mapping = {}

    asset_ids = [str(asset.get("id") or "") for asset in draft_assets(draft_document, asset_type)]
    extra_asset_ids = [asset_id for asset_id in asset_mapping if asset_id not in asset_ids]
    return [(asset_id, mapped_column(asset_mapping, asset_id)) for asset_id in asset_ids + extra_asset_ids]


def parse_timestamp(value: str) -> datetime:
    if not value:
        raise ValueError("empty timestamp")
    return datetime.fromisoformat(value.replace("Z", "+00:00"))


def parse_mapped_float(
    row: dict[str, str],
    column: str,
    row_index: int,
    label: str,
    errors: list[str],
) -> float | None:
    value = str(row.get(column) or "").strip()
    try:
        number = float(value)
    except ValueError:
        errors.append(f"row {row_index}: {label} must be numeric")
        return None
    if not math.isfinite(number):
        errors.append(f"row {row_index}: {label} must be finite")
        return None
    return number


def parse_mapped_float_or_zero(
    row: dict[str, str],
    column: str | None,
    row_index: int,
    label: str,
    errors: list[str],
) -> float | None:
    if not column:
        return 0.0
    return parse_mapped_float(row, column, row_index, label, errors)


def suggest_mappings(columns: list[str], draft_document: dict[str, Any]) -> dict[str, Any]:
    normalized_columns = {normalize_column(column): column for column in columns}
    suggestions: dict[str, Any] = {
        "timestamp": find_first_column(
            normalized_columns,
            ["timestamp", "datetime", "date_time", "period_start", "time"],
        ),
        "duration_hours": find_first_column(
            normalized_columns,
            ["duration_hours", "duration_hour", "duration_h", "duration", "hours"],
        ),
        "price_usd_per_mwh": find_first_column(
            normalized_columns,
            ["price_usd_per_mwh", "legacy_price_usd_per_mwh", "price", "market_price"],
        ),
        "import_price_usd_per_mwh": find_first_column(
            normalized_columns,
            ["import_price_usd_per_mwh", "import_price", "buy_price_usd_per_mwh", "buy_price"],
        ),
        "export_price_usd_per_mwh": find_first_column(
            normalized_columns,
            ["export_price_usd_per_mwh", "export_price", "sell_price_usd_per_mwh", "sell_price"],
        ),
        "renewable_available_power_mw": {},
        "load_demand_mw": {},
        "hydro_inflow_m3s": {},
    }

    for asset in draft_assets(draft_document, "renewable"):
        asset_id = str(asset.get("id") or "")
        matched = find_asset_column(
            columns,
            asset_id,
            ["available", "availability", "available_mw", "renewable_available_power_mw", "generation"],
        )
        if matched:
            suggestions["renewable_available_power_mw"][asset_id] = matched

    for asset in draft_assets(draft_document, "load"):
        asset_id = str(asset.get("id") or "")
        matched = find_asset_column(
            columns,
            asset_id,
            ["demand", "demand_mw", "load", "load_mw", "load_demand_mw"],
        )
        if matched:
            suggestions["load_demand_mw"][asset_id] = matched

    hydro_assets = draft_assets(draft_document, "hydro")
    generic_hydro_inflow = None
    if len(hydro_assets) == 1:
        generic_hydro_inflow = find_first_column(
            normalized_columns,
            ["hydro_inflow_m3s", "inflow_m3s", "inflow"],
        )

    for asset in hydro_assets:
        asset_id = str(asset.get("id") or "")
        matched = find_asset_column(
            columns,
            asset_id,
            ["inflow", "inflow_m3s", "hydro_inflow_m3s"],
        ) or generic_hydro_inflow
        if matched:
            suggestions["hydro_inflow_m3s"][asset_id] = matched

    return suggestions


def draft_assets(document: dict[str, Any], asset_type: str) -> list[dict[str, Any]]:
    assets = document.get("assets")
    if not isinstance(assets, list):
        return []
    return [asset for asset in assets if isinstance(asset, dict) and asset.get("type") == asset_type]


def find_first_column(normalized_columns: dict[str, str], aliases: list[str]) -> str | None:
    for alias in aliases:
        column = normalized_columns.get(normalize_column(alias))
        if column:
            return column
    return None


def find_asset_column(columns: list[str], asset_id: str, suffixes: list[str]) -> str | None:
    normalized_asset_id = normalize_column(asset_id)
    for column in columns:
        normalized = normalize_column(column)
        if normalized == normalized_asset_id:
            return column
        if normalized.startswith(f"{normalized_asset_id}_"):
            tail = normalized[len(normalized_asset_id) + 1 :]
            if tail in {normalize_column(suffix) for suffix in suffixes}:
                return column
        if normalized.endswith(f"_{normalized_asset_id}"):
            head = normalized[: -len(normalized_asset_id) - 1]
            if head in {normalize_column(suffix) for suffix in suffixes}:
                return column
    return None


def normalize_column(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")
